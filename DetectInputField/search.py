from requests.api import request
from requests.exceptions import SSLError
from requests_html import HTMLSession
import requests
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pyppeteer


WORD_LIST = []
FILE_NAME = 'webpages_inputdata.xlsx'
wb = load_workbook(FILE_NAME)


# Appending keywords to the list from excel sheet
def get_keywords():
    keywords = wb['Keywords']
    for row in range(2, keywords.max_row + 1):
        # generates the links one by one
        if value := keywords[f"A{row}"].value:
                WORD_LIST.append(value)


# Returns True if any of the given texts is found
def find_text(response, word_list):
    html = response.html.html
    found = None
    
    for word in word_list:
        result = html.lower().find(word.lower())
        if not result == -1:
            found = word
            break
    
    return found


# finds input labels and its text
def find_input_labels(response):
    html = response.html
    labels = html.find('label')
    for label in labels:
        print(label.text)


def find_inputs(response):
    html = response.html
    inputs = html.find('input')
    visible_inputs = [inp for inp in inputs if 'type' in inp.attrs and inp.attrs['type'] != 'hidden']
    return len(visible_inputs) > 1




# Prepares the excel sheets and names the columns
def customize_excel_sheet():
    global wb
    output = wb.create_sheet('Output') if 'Output' not in wb.sheetnames else wb['Output']
    errors = wb.create_sheet('Errors') if 'Errors' not in wb.sheetnames else wb['Errors']
    
    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor='E8E8E8', fill_type='solid')

    # editing the output sheet
    output_column = zip(('A',  'B', 'C'), ('URL', 'Can Input', 'Keyword Found'))
    for col, value in output_column:
        cell = output[f'{col}1']
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        output.freeze_panes = cell

        # fixing the column width
        output.column_dimensions[col].width = 20


# Generates the input links
def generate_input_urls():
    global wb
    inputs = wb['Input']
    for row in range(2, inputs.max_row + 1):
        # generates the links one by one
        if value := inputs[f"A{row}"].value:
            yield value


# corrects the url
def correct_url(url, session):
    global wb
    errors = wb['Errors']
    if not url.startswith('http'):
        url = 'https://' + url
    try:
        r = session.get(url)
    except SSLError:
        r = session.get(url, verify=False)
    except Exception as e:
        r = None
        errors.append((url, str(e)))
    return r if r and r.ok else None


# Returns data in a structured format
def get_data(response):
    keyword_found = find_text(response, WORD_LIST)
    can_input = find_inputs(response) or bool(keyword_found)
    data = {'can_input': can_input, 'keyword_found': keyword_found}
    return data

    # THIS BLOCK OF CODE WILL BE APPLICABLE IF KEYWORDS ARE GROUPED BY TYPES
    # keyword = []
    # for group in keyword_groups:
    #     found = find_text(response, WORD_LIST)
    #     if found:
    #         keyword.append((group, found))


def insert_data_to_excel():
    global wb
    # gets keywords from excel sheet and appends it to the list
    get_keywords() 
    customize_excel_sheet()
    session = HTMLSession()
    output = wb['Output']
    errors = wb['Errors']

    # iterating through the input urls
    for url in generate_input_urls():
        if response := correct_url(url, session):
            print("[PROCESSING] - ", f'[{url}]', end=' ')
            try:
                response.html.render(timeout=100)
                data = get_data(response)
                # appending data to the excel sheet
                output.append((
                    url,
                    data['can_input'],
                    data['keyword_found'],
                ))
                print('[Successful]')
            except pyppeteer.errors.TimeoutError as e:
                errors.append((url, str(e)))
                print(['[Timeout Error]'])
    wb.save(FILE_NAME)

insert_data_to_excel()


