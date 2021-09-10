from bs4 import BeautifulSoup
from collections import deque
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
import requests
from requests import Session
from termcolor import colored
import random

# determines if links should be case sensitive or not
CASE_SENSITIVE = False

HEADERS_LIST = [
'Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13',
'Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko',
'Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201',
'Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16',
'Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre'
]


def init_driver():
    """Initialize the chrome web driver with necessary parameters and return the driver"""
    # set up the chrome options
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--log-level=3")

    # initialize the chrome web driver
    driver = webdriver.Chrome(chrome_options=chrome_options)

    return driver


# get html using selenium
def get_html(driver: webdriver.Chrome, url: str, sleep=2) -> str:
    """Use selenium to run javascript and get the inner html of the webpage"""
    url = "https://" + url if not url.startswith("http") else url

    try:
        driver.get(url)
        # wait sometime for the page to load
        time.sleep(sleep)

        # run javascript to get the innerHTML
        html = driver.page_source

        # print a 'scraping' message
        print(colored(f"Successfully Scraped {url}", 'green'))
        return html
    except Exception as exception:
        print(colored(f"{url} failed due to {str(exception)}", 'red'))


# Run get_html and return a soup object
def get_soup(driver: webdriver.Chrome, url: str) -> BeautifulSoup:
    """Get the response and return a BeautifulSoup object"""
    html = get_html(driver, url)
    if html:
        soup = BeautifulSoup(html, "html.parser")

        return soup


def website_urls_generator(websites, start=2):
    for row in range(start, websites.max_row + 1):
        cell = websites[f'A{row}']
        if cell.value:
            yield cell.value 



# takes an url and returns its status code and final redirected url
def get_final_link(url, session, wb):
    if url:
        # creating the request Session
        header = {
            "User-Agent": random.choice(HEADERS_LIST),
            "X-Requested-With": "XMLHttpRequest",
        }
        session.headers.update(header)

        url = "https://" + url if not url.startswith("http") else url
        final_url = url
        try:
            response = session.get(url)
            final_url = response.url
        except requests.exceptions.SSLError:
            response = session.get(url, verify=False)
            final_url = response.url
        except requests.exceptions.ConnectionError as err:
            final_url = url
            error = str(err)
            wb["Errors"].append(
                ("Failed to get the final url for ", url, "Due to", error)
            )
        except Exception as err:
            wb["Errors"].append(
                ("Failed to get the final url for ", url, "Due to", str(err))
            )
        return final_url


def map_links(driver, main_url):
    main_url = 'https://' + main_url if not main_url.startswith('http') else main_url
    main_url = main_url.lower() if not CASE_SENSITIVE else main_url

    # a queue of urls to be crawled next
    new_urls = deque([main_url])
    # a set of urls that we have already processed 
    processed_urls = set()

    while len(new_urls):    
        # move url from the queue to processed url set    
        url = new_urls.popleft()    
        processed_urls.add(url)    

        # do not find urls on external links, just yield them and go on to the next 
        if url.startswith('http') and not url.startswith(main_url):
            yield url
            continue
            

        # extract base url to resolve relative links
        base = main_url  # base = parts.netloc
        strip_base = base.replace("www.", "")
        base_url = main_url + '/' if not main_url.endswith('/') else main_url
        extra_in_base = base_url.split('/')[-2]
        path = main_url

        # initialize beautifulSoup to extract links from html document
        soup = get_soup(driver, url)
        if soup:
            for link in soup.find_all('a'):
                # extract link url from the anchor    
                anchor = link.attrs["href"] if 'href' in link.attrs else ''
                anchor = anchor.lower() if not CASE_SENSITIVE else anchor
                # stopping duplication and filtering out usable links

                # checking if the url is a image url
                media_extentions = ['.jpg', '.jpeg', '.gif', '.png', 'bmp', 'svg', 'mp4', 'wmv', 'mp3', 'pdf']
                media_url = False
                for extention in media_extentions:
                    if extention in anchor.lower():
                        media_url = True
                # checking if url is usable
                conditions = (not anchor.endswith('#')) and (anchor.count('#') <= 1) and ('mailto' not in anchor) and ('tel' not in anchor)

                if conditions and not media_url:
                    if anchor.startswith(main_url):
                        if not anchor in new_urls and not anchor in processed_urls:
                            new_urls.append(anchor)
                    elif anchor.startswith(f'/{extra_in_base}'):
                        local_link = base_url + anchor.lstrip(f'/{extra_in_base}') 
                        if not local_link in new_urls and not local_link in processed_urls:
                            new_urls.append(local_link)
                    elif strip_base in anchor: 
                        if not anchor in new_urls and not anchor in processed_urls:
                            new_urls.append(anchor)
                    elif not anchor.startswith('http'):  
                        local_link = path + anchor 
                        if not local_link in new_urls and not local_link in processed_urls:
                            new_urls.append(local_link)
                    # add external links to the list
                    elif not anchor.startswith(main_url):
                        if not anchor in new_urls and not anchor in processed_urls:
                            new_urls.append(anchor) 
            yield url


def map_homepage_links(driver, main_url):
    "Return all links in the given url"
    main_url = 'https://' + main_url if not main_url.startswith('http') else main_url
    main_url = main_url.lower() if not CASE_SENSITIVE else main_url

    soup = get_soup(driver, main_url)
    processed_urls = set()

    # extract base url to resolve relative links
    base = main_url  # base = parts.netloc
    strip_base = base.replace("www.", "")
    base_url = main_url + '/' if not main_url.endswith('/') else main_url
    extra_in_base = base_url.split('/')[-2]
    path = main_url

    if soup:
        for link in soup.find_all('a'):
            # extract link url from the anchor    
            anchor = link.attrs["href"] if 'href' in link.attrs else ''
            anchor = anchor.lower() if not CASE_SENSITIVE else anchor
            # stopping duplication and filtering out usable links
            media_extentions = ['.jpg', '.jpeg', '.gif', '.png', 'bmp', 'svg', 'mp4', 'wmv', 'mp3', 'pdf']
            media_url = False
            for extention in media_extentions:
                if extention in anchor.lower():
                    media_url = True
            # checking if url is usable
            conditions = (not anchor.endswith('#')) and (anchor.count('#') <= 1) and ('mailto' not in anchor) and ('tel' not in anchor)

            if conditions and not media_url:
                if anchor.startswith(main_url):
                    if anchor not in processed_urls:
                        processed_urls.add(anchor)
                        yield anchor

                elif anchor.startswith(f'/{extra_in_base}'):
                    local_link = base_url + anchor.lstrip(f'/{extra_in_base}') 
                    if local_link not in processed_urls:
                        processed_urls.add(local_link)
                        yield local_link

                elif strip_base in anchor: 
                    if anchor not in processed_urls:
                        processed_urls.add(anchor)
                        yield anchor

                elif not anchor.startswith('http'):  
                    local_link = path + anchor 
                    if local_link not in processed_urls:
                        processed_urls.add(local_link)
                        yield local_link

                # yield external links too
                elif not anchor.startswith(main_url):
                    if anchor not in processed_urls:
                        processed_urls.add(anchor)
                        yield anchor


# combine all the functions to get the expected output
def main(file_path, driver, homepage_only=True, start=2):
    wb = load_workbook(file_path)
    websites = wb['Websites']
    webpages = wb.create_sheet('Webpages') if 'Webpages' not in wb.sheetnames else wb['Webpages']
    errors = wb.create_sheet('Errors') if 'Errors' not in wb.sheetnames else wb['Errors']

    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor='E8E8E8', fill_type='solid')

    # editing the users sheet
    webpage_columns = zip(('A',  'B', 'C'), ('Website', 'Webpage', 'Final Link Destination'))
    for col, value in webpage_columns:
        cell = webpages[f'{col}1']
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        webpages.freeze_panes = cell

        # fixing the column width
        webpages.column_dimensions[col].width = 20
    
    # decide on homepage only or to scrape whole website
    map_urls = (lambda homepage_only: map_homepage_links if homepage_only else map_links)(homepage_only)

    # iterate over all website urls and map the urls 
    for website in website_urls_generator(websites, start=start):
        count = 0
        session = Session()

        for webpage in map_urls(driver, website):
            count += 1
            # get the final destination of a link
            final_link = get_final_link(webpage, session, wb)

            # print a colored feedback of the process
            print(f"{colored(website, 'magenta')}({count}) --> {colored(webpage, 'blue')} --> {colored(final_link, 'blue')}")

            # add the data to excel
            webpages.append((
                website, webpage, final_link
            ))

            # save after each link is added
            wb.save(FILE_PATH)

    # save at the end of the program
    wb.save(FILE_PATH)
    
    print(f"Saved the urls into {FILE_PATH}")

    # close the driver
    driver.close()


if __name__ == '__main__':
    FILE_PATH = 'webpages.xlsx'
    HOMEPAGE_ONLY = True
    NEW_URL_STARTING_ROW = 2
    driver = init_driver()
    main(FILE_PATH, driver,homepage_only=HOMEPAGE_ONLY, start=NEW_URL_STARTING_ROW)

    