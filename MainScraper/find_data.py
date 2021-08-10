import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from list_urls import map_urls, get_soup, init_driver
from termcolor import colored




def find_gtm(soup):
    """ "Gets the GTM from a web page"""
    if soup and soup.find("head"):
        gtm_container = soup.head.findAll("script", text=re.compile(r"GTM"))
        if gtm_container:
            gtm_code = re.search("GTM-[A-Z0-9]{6,7}", str(gtm_container))
            gtm_code = (
                re.search("GTM-[A-Z0-9]{6,7}", str(gtm_container))[0]
                if gtm_code
                else None
            )

            return gtm_code


def find_statement(soup, statement):
    """ "Checks for the presence for the given statement"""
    statement_list = soup.findAll(text=re.compile(f"{statement}"))
    return bool(statement_list)


def find_link(soup, link):
    """ "Checks for the presence for the given link"""
    a_tags = soup.findAll("a")
    link = str(link).rstrip("/") if link and link.endswith("/") else str(link)
    for a in a_tags:
        href = a.attrs["href"] if "href" in a.attrs else ""
        if link.lower() in href.lower():
            return True
    return False


def find_html_lang(soup):
    """ "Gets the language of a webpage if mentioned in html tag"""
    html = soup.find("html")
    if html and "lang" in html.attrs:
        return html.attrs["lang"]


def get_data(
    driver, url, statement1, statement2, statement3, link1, link2, link3, link4, link5, link6
):
    """Gets the gtm and lang from a page and check for presence of links and statements mentioned."""

    for i in map_urls(url):
        soup = get_soup(driver, url)

        gtm = find_gtm(soup)
        statement_1 = find_statement(soup, statement1)
        statement_2 = find_statement(soup, statement2)
        statement_3 = find_statement(soup, statement3)
        link_1 = find_link(soup, link1)
        link_2 = find_link(soup, link2)
        link_3 = find_link(soup, link3)
        link_4 = find_link(soup, link4)
        link_5 = find_link(soup, link5)
        link_6 = find_link(soup, link6)
        lang = find_html_lang(soup)
        no_direct_yt_link = find_link(soup, "youtube.com")

        data_dict = {
            "url": i,
            "GTM": gtm,
            "statement1": statement_1,
            "statement2": statement_2,
            "statement3": statement_3,
            "link1": link_1,
            "link2": link_2,
            "link3": link_3,
            "link4": link_4,
            "link5": link_5,
            "link6": link_6,
            "lang": lang,
            "no_direct_yt_link": no_direct_yt_link,
        }

        yield data_dict


def get_homepage_data(
    driver, url, statement1, statement2, statement3, link1, link2, link3, link4, link5, link6
):
    """Gets the gtm and lang from the  homepage and check for presence of links and statements mentioned."""
    soup = get_soup(driver, url)

    gtm = find_gtm(soup)
    statement_1 = find_statement(soup, statement1)
    statement_2 = find_statement(soup, statement2)
    statement_3 = find_statement(soup, statement3)
    link_1 = find_link(soup, link1)
    link_2 = find_link(soup, link2)
    link_3 = find_link(soup, link3)
    link_4 = find_link(soup, link4)
    link_5 = find_link(soup, link5)
    link_6 = find_link(soup, link6)
    lang = find_html_lang(soup)
    no_direct_yt_link = not (
        find_link(soup, "youtube.com/em") or find_link(soup, "youtube.com/wa")
    )

    data_dict = {
        "url": url,
        "GTM": gtm,
        "statement1": statement_1,
        "statement2": statement_2,
        "statement3": statement_3,
        "link1": link_1,
        "link2": link_2,
        "link3": link_3,
        "link4": link_4,
        "link5": link_5,
        "link6": link_6,
        "lang": lang,
        "no_direct_yt_link": no_direct_yt_link,
    }

    return data_dict


def main(FILE_PATH):
    HOMEPAGE_ONLY = True
    NEW_URL_STARTING_ROW = 2
    driver = init_driver()


    wb = load_workbook(FILE_PATH)
    websites = wb["Websites"]
    page_results = (
        wb.create_sheet("Page Results")
        if "Page Results" not in wb.sheetnames
        else wb["Page Results"]
    )
    errors = (
        wb.create_sheet("Errors") if "Errors" not in wb.sheetnames else wb["Errors"]
    )

    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor="E8E8E8", fill_type="solid")

    # editing the users sheet
    page_results_columns = zip(
        ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"),
        (
            "Webpage",
            "GTM",
            "href-lang",
            "link1",
            "link2",
            "link3",
            "link4",
            "link5",
            "link6",
            "statement1",
            "statement2",
            "statement3",
            "Page contains no direct youtube links",
        ),
    )
    for col, value in page_results_columns:
        cell = page_results[f"{col}1"]
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        page_results.freeze_panes = cell

        # fixing the column width
        page_results.column_dimensions[col].width = 20

    # Editing the errors sheet
    for col, value in (("A", "Error URL"), ("B", "Error Name")):
        cell = errors[f"{col}1"]
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        # fixing the column width
        errors.column_dimensions[col].width = 40

    def website_urls_generator():
        start = NEW_URL_STARTING_ROW
        for row in range(start, websites.max_row + 1):
            website = websites[f"A{row}"].value
            link1 = websites[f"B{row}"].value
            link2 = websites[f"C{row}"].value
            link3 = websites[f"D{row}"].value
            link4 = websites[f"E{row}"].value
            link5 = websites[f"F{row}"].value
            link6 = websites[f"G{row}"].value
            statement1 = websites[f"H{row}"].value
            statement2 = websites[f"I{row}"].value
            statement3 = websites[f"J{row}"].value

            if website:
                data_row = {
                    "url": website,
                    "link1": link1,
                    "link2": link2,
                    "link3": link3,
                    "link4": link4,
                    "link5": link5,
                    "link6": link6,
                    "statement1": statement1,
                    "statement2": statement2,
                    "statement3": statement3,
                }
                yield data_row

    for website in website_urls_generator():
        if HOMEPAGE_ONLY:
            # fetching data from the excel sheet
            try:
                data_dict = get_homepage_data(
                    driver,
                    website["url"],
                    website["statement1"],
                    website["statement2"],
                    website["statement3"],
                    website["link1"],
                    website["link2"],
                    website["link3"],
                    website["link4"],
                    website["link5"],
                    website["link6"],
                )

                # appending data to the excel sheet
                page_results.append(
                    (
                        data_dict["url"],
                        data_dict["GTM"],
                        data_dict["lang"],
                        data_dict["link1"],
                        data_dict["link2"],
                        data_dict["link3"],
                        data_dict["link4"],
                        data_dict["link5"],
                        data_dict["link6"],
                        data_dict["statement1"],
                        data_dict["statement2"],
                        data_dict["statement3"],
                        data_dict["no_direct_yt_link"],
                    )
                )
                # save after scraping each site
                wb.save(FILE_PATH)
            except Exception as e:
                print(colored(f"{website['url']} Fails due to {e}", 'red'))
                errors.append((website["url"], str(e)))

        else:
            for webpage in get_data(
                driver,
                website["url"],
                website["statement1"],
                website["statement2"],
                website["statement3"],
                website["link1"],
                website["link2"],
                website["link3"],
                website["link4"],
                website["link5"],
                website["link6"],
            ):
                try:
                    page_results.append(
                        (
                            webpage["url"],
                            webpage["GTM"],
                            webpage["lang"],
                            webpage["link1"],
                            webpage["link2"],
                            webpage["link3"],
                            webpage["link4"],
                            webpage["link5"],
                            webpage["link6"],
                            webpage["statement1"],
                            webpage["statement2"],
                            webpage["statement3"],
                            webpage["no_direct_yt_link"],
                        )
                    )
                    # save after scraping each site
                    wb.save(FILE_PATH)
                except Exception as e:
                    print(colored(f"{website['url']} Fails due to {e}", 'red'))
                    errors.append((webpage, str(e)))

    try:
        wb.save(FILE_PATH)
        print(f'Saved the data into {FILE_PATH}')
    except PermissionError:
        print(colored("PermissionError: Please close the working excel file", 'red'))


if __name__ == "__main__":
    FILE_PATH = "webpages.xlsx"  
    main(FILE_PATH)

