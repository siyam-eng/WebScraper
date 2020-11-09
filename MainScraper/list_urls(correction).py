from bs4 import BeautifulSoup
import requests
import requests.exceptions
from urllib.parse import urlsplit
from collections import deque
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import random

# determines if links should be case sensitive or not
CASE_SENSITIVE = False

def map_urls(main_url):
    main_url = main_url.lower() if not CASE_SENSITIVE else main_url
    HEADERS_LIST = [
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13',
    'Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201',
    'Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16',
    'Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre'
    ]
    try:    
        session = requests.Session()
        session.get(main_url)
        HEADER = {'User-Agent': random.choice(HEADERS_LIST), 'X-Requested-With': 'XMLHttpRequest'}
        session.headers.update(HEADER)
    except(requests.exceptions.MissingSchema,
            requests.exceptions.ConnectionError, 
            requests.exceptions.InvalidURL, 
            requests.exceptions.InvalidSchema, 
            requests.exceptions.TooManyRedirects, 
            Exception,
        ):    
        pass
    # a queue of urls to be crawled next
    new_urls = deque([main_url])
    # a set of urls that we have already processed 
    processed_urls = set()

    while len(new_urls):    
        # move url from the queue to processed url set    
        url = new_urls.popleft()    
        processed_urls.add(url)    

        try:    
            response = session.get(url)
        except(requests.exceptions.MissingSchema,
               requests.exceptions.ConnectionError, 
               requests.exceptions.InvalidURL, 
               requests.exceptions.InvalidSchema, 
               requests.exceptions.TooManyRedirects, 
               Exception,
               ):    
            # add broken urls to it’s own set, then continue       
            continue
        # extract base url to resolve relative links
        parts = urlsplit(url)
        base = main_url  # base = parts.netloc
        strip_base = base.replace("www.", "")
        base_url = main_url + '/' if not main_url.endswith('/') else main_url
        extra_in_base = base_url.split('/')[-2]
        path = main_url

        # initialize beautifulSoup to extract links from html document
        soup = BeautifulSoup(response.content, 'html.parser')
        for link in soup.find_all('a'):
            # extract link url from the anchor    
            anchor = link.attrs["href"] if 'href' in link.attrs else ''
            anchor = anchor.lower() if not CASE_SENSITIVE else anchor
            # stopping duplication and filtering out usable links

            # checking if the url is a image url
            media_extentions = ['.jpg', '.jpeg', '.gif', '.png', 'bmp', 'svg', 'mp4', 'wmv', 'mp3', 'pdf']
            media_url = False
            for extention in media_extentions:
                if extention in anchor.lower():
                    media_url = True
            # checking if url is usable
            conditions = (not anchor.endswith('#')) and (anchor.count('#') <= 1) and ('mailto' not in anchor) and ('tel' not in anchor)

            if conditions and not media_url:
                if anchor.startswith(main_url):
                    if not anchor in new_urls and not anchor in processed_urls:
                        new_urls.append(anchor)
                elif anchor.startswith(f'/{extra_in_base}'):
                    local_link = base_url + anchor.lstrip(f'/{extra_in_base}') 
                    if not local_link in new_urls and not local_link in processed_urls:
                        new_urls.append(local_link)
                elif strip_base in anchor: 
                    if not anchor in new_urls and not anchor in processed_urls:
                        new_urls.append(anchor)
                elif not anchor.startswith('http'):  
                    local_link = path + anchor 
                    if not local_link in new_urls and not local_link in processed_urls:
                        new_urls.append(local_link)

        yield url


if __name__ == '__main__':
    FILE_PATH = 'webpages.xlsx'
    NEW_URL_STARTING_ROW = 2

    wb = load_workbook(FILE_PATH)
    websites = wb['Websites']
    webpages = wb.create_sheet('Webpages') if 'Webpages' not in wb.sheetnames else wb['Webpages']

    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor='E8E8E8', fill_type='solid')

    # editing the users sheet
    webpage_columns = zip(('A',  'B'), ('Website', 'Webpage'))
    for col, value in webpage_columns:
        cell = webpages[f'{col}1']
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        webpages.freeze_panes = cell

        # fixing the column width
        webpages.column_dimensions[col].width = 20

    def website_urls_generator():
        start = NEW_URL_STARTING_ROW
        for row in range(start, websites.max_row + 1):
            cell = websites[f'A{row}']
            if cell.value:
                yield cell.value 


    for website in website_urls_generator():
        for webpage in map_urls(website):
            print(webpage)
            webpages.append((
                website, webpage
            ))
    wb.save(FILE_PATH)
    

