import requests
import requests.exceptions

import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import random
from list_urls import get_soup, init_driver, website_urls_generator
from termcolor import colored


HEADERS_LIST = [
'Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13',
'Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko',
'Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201',
'Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16',
'Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre'
]

def find_indices(str, ch):
    """Find all indices of a character in a string"""
    indexes = []
    for i, ltr in enumerate(str):
        if ltr == ch:
            indexes.append(i)
    return indexes 


def find_code(soup, code):
    """Finds all codes matching with the given code"""
    codes = soup.findAll(text=re.compile(f'{code}'))
    codes2 = []
    for c in codes:
        code_list = re.findall(f'[A-Z-/]*{code}[A-Z0-9a-z-/]*', str(c))
        for code2 in code_list:
            code2 = code2.lstrip('-') if code2.startswith('-') else code2
            code2 = code2.lstrip('/') if code2.startswith('/') else code2
            if '/' in code2 or '-' in code2:
                # print(code2)
                # limiting the lenth of the code upto 3rd seperator
                if code2.count('-') > 4:
                    third_hyphen = find_indices(code2, '-')[4]
                    code2 = code2[: third_hyphen]
                if code2.count('/') > 3:
                    third_slash = find_indices(code2, '/')[3]
                    code2 = code2[: third_slash]
                if code2.count('/') > 2 and '-' in code2:
                    code2 = code2[: code2.index('-')]
                
                codes2.append(code2)
    return codes2


def get_codes(url, soup, code_generator):
    for code in code_generator:
        # codes containing present code on page
        present_code = find_code(soup, code)
        for code in present_code:
            code_type = '/' if '/' in code else '-'
            yield {'url': url, 'code': code, 'type': code_type, 'length': len(code)}


def codes_lookups_generator(code_lookups, start=2):
    for row in range(start, code_lookups.max_row + 1):
        cell = code_lookups[f'A{row}']
        if cell.value:
            yield cell.value 


def get_response(url):
    """Send a http request to the given url and return the response"""
    header = {'User-Agent': random.choice(HEADERS_LIST), 'X-Requested-With': 'XMLHttpRequest'}
    url = 'https://' + url if not url.startswith('http') else url
    try:
        response = requests.get(url, headers=header)
        if not response.ok:
            response = requests.get(url)
        
        return response

    except Exception as exception:
        print(colored(f"{url} failed due to {exception}", 'red'))



def main(file_path):
    driver = init_driver()

    # initialize the necessary excel sheets
    wb = load_workbook(file_path)
    webpages = wb['Websites']
    code_lookups = wb['Code_Lookups']
    codes = wb.create_sheet('Codes') if 'Codes' not in wb.sheetnames else wb['Codes']

    # specify the styles
    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor='E8E8E8', fill_type='solid')

    # edit the users sheet
    codes_columns = zip(('A',  'B', 'C', 'D'), ('Web Page', 'Code Type', 'Code', 'Lenth'))
    for col, value in codes_columns:
        cell = codes[f'{col}1']
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        codes.freeze_panes = cell

        # fix the column width
        codes.column_dimensions[col].width = 20

    # iterate over the webpages given on the excel file
    for webpage in website_urls_generator(webpages):
        soup = get_soup(driver, webpage)

        if soup:
            code_generator = codes_lookups_generator(code_lookups)
            for code in get_codes(webpage, soup, code_generator):
                print(colored(code, 'magenta'))
                codes.append((
                    code['url'],
                    code['code'],
                    code['type'],
                    code['length'],
                ))
                # save after every code
                wb.save(file_path)
    # save at the end
    wb.save(file_path)
    print(f"Saved the CODES into {file_path}")

    # close the driver
    driver.close()


if __name__ == '__main__':
    FILE_PATH = 'webpages.xlsx'
    main(FILE_PATH)
